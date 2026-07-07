from pathlib import Path
import csv
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FILES = {
    "brand": Path(r"C:/Users/leiss/Downloads/brand-export.D72zJqPJQ5hmke3OOvojugWmZ8NPQmHm05o4YlEpHZa3XYj932uQlkWrq1ZdcdKZZlX62qsOrPm3G.xlsx"),
    "ssmt": Path(r"C:/Users/leiss/Downloads/SEA Standard Menu Template.xlsx"),
    "menuworks": Path(r"C:/Users/leiss/Downloads/MenuWorks_Menu_Item_bd043994-d2ba-4654-94de-692575c178a0.csv"),
}

SAMPLE_NAMES = {
    "Kachumbar",
    "Mango Sticky Rice",
    "Turkey Apple Brie Sandwich with Apple",
    "Turkey Burger Lettuce Wrap with Quinoa Salad",
    "Turkey Burger on Lettuce Wrap with Orzo Salad",
    "Turkey Swiss Flatbread Sandwich with Apple",
    "Turkey Swiss Flatbread Sandwich with Tabouli Salad",
    "Smashburger",
    "Smashburger Patty",
    "Spicy Firebird Sandwich",
    "Portobello Tofu Teriyaki",
    "Teriyaki Salad",
    "Green Curry Pork Bowl",
    "Green Curry Tofu Bowl",
}


def cell_text(cell):
    if cell.value is None:
        return ""
    return str(cell.value).strip()


def inspect_csv(path):
    try:
        with path.open("r", encoding="utf-8-sig", errors="strict", newline="") as handle:
            handle.read(4096)
        encoding = "utf-8-sig"
    except UnicodeDecodeError:
        encoding = "cp1252"

    with path.open("r", encoding=encoding, errors="replace", newline="") as handle:
        reader = csv.DictReader(handle)
        headers = reader.fieldnames or []
        print(f"MENUWORKS encoding={encoding} header_count={len(headers)}")
        for index, header in enumerate(headers, start=1):
            lower = header.lower()
            if any(token in lower for token in ["mrn", "recipe", "menu", "station", "category", "description", "item", "price", "calorie"]):
                print(f"  {index}: {header}")

        print("MENUWORKS sample MRNs")
        for row in reader:
            joined = " | ".join(str(row.get(header, "")) for header in headers[:18])
            if any(name.lower() in joined.lower() for name in SAMPLE_NAMES):
                values = {
                    header: row.get(header, "")
                    for header in headers
                    if any(token in header.lower() for token in ["mrn", "recipe", "menu", "item", "price", "category"])
                }
                print(values)


def print_sheet_cells(sheet, rows=8, max_column=30):
    for row_number in range(1, min(sheet.max_row, rows) + 1):
        cells = []
        for column_number in range(1, min(sheet.max_column, max_column) + 1):
            cell = sheet.cell(row_number, column_number)
            text = cell_text(cell)
            if text:
                cells.append(f"{get_column_letter(column_number)}={text} [{cell.data_type}/{cell.number_format}]")
        if cells:
            print(f"  row {row_number}: " + " ; ".join(cells))


def inspect_brand(path):
    workbook = load_workbook(path, read_only=False, data_only=False)
    print(f"BRAND sheets={workbook.sheetnames}")
    for sheet_name in ["Brand", "Items", "Modifiers"]:
        if sheet_name not in workbook.sheetnames:
            print(f"{sheet_name}: MISSING")
            continue
        sheet = workbook[sheet_name]
        print(f"{sheet_name}: rows={sheet.max_row} columns={sheet.max_column}")
        if sheet_name == "Items":
            print_sheet_cells(sheet, rows=8, max_column=26)
            for row_number in range(1, min(sheet.max_row, 8) + 1):
                selected = []
                for col in ["D", "E", "F", "G", "Z"]:
                    cell = sheet[f"{col}{row_number}"]
                    selected.append(f"{col}{row_number}={cell_text(cell)} [{cell.data_type}/{cell.number_format}]")
                print("  item-map " + " ; ".join(selected))
        elif sheet_name == "Modifiers":
            print_sheet_cells(sheet, rows=10, max_column=52)
            for row_number in range(1, min(sheet.max_row, 12) + 1):
                selected = []
                for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "AZ"]:
                    cell = sheet[f"{col}{row_number}"]
                    selected.append(f"{col}{row_number}={cell_text(cell)} [{cell.data_type}/{cell.number_format}]")
                print("  modifier-map " + " ; ".join(selected))
        else:
            print_sheet_cells(sheet, rows=12, max_column=12)


def inspect_ssmt(path):
    workbook = load_workbook(path, read_only=False, data_only=False)
    print(f"SSMT sheets={workbook.sheetnames[:20]} total={len(workbook.sheetnames)}")
    for sheet in workbook.worksheets[:12]:
        candidates = []
        for row_number in range(1, min(sheet.max_row, 25) + 1):
            row_text = " | ".join(cell_text(sheet.cell(row_number, col)) for col in range(1, min(sheet.max_column, 20) + 1))
            if any(token in row_text.lower() for token in ["mrn", "recipe", "menu item", "description", "category", "remove"]):
                candidates.append((row_number, row_text))
        if candidates:
            print(f"SSMT candidate sheet={sheet.title} rows={sheet.max_row} cols={sheet.max_column}")
            for row_number, row_text in candidates[:5]:
                print(f"  row {row_number}: {row_text}")


for name, path in FILES.items():
    print(f"FILE {name} exists={path.exists()} size={path.stat().st_size if path.exists() else None}")

inspect_csv(FILES["menuworks"])
inspect_brand(FILES["brand"])
inspect_ssmt(FILES["ssmt"])
